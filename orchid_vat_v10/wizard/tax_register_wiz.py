from odoo import models, fields, api
from odoo.tools.translate import _
from odoo.exceptions import UserError, AccessError

import time
from datetime import date
from datetime import datetime,date
from datetime import timedelta

import base64
import xlwt
from io import StringIO
from io import BytesIO

from pprint import pprint
import logging
from odoo import tools
from openpyxl.workbook import Workbook

_logger = logging.getLogger(__name__)
from collections import defaultdict


class OrchidTaxRegister(models.TransientModel):
	"""
	Orchid Tax Register .
	"""
	_name = 'orchid.tax.register'
	_description = 'Tax Register '

	from_date = fields.Date(string='From Date')
	to_date = fields.Date(string='To Date')
	excel_file = fields.Binary(string='Excel Report',readonly=True)
	file_name = fields.Char(string='Excel File',readonly=True)
	od_product = fields.Boolean(string='Product Wise')
	journal_id = fields.Many2many('account.journal',string='Journals')
	account_id = fields.Many2many('account.account',string='Accounts')

	def generate(self):
		if self.from_date > self.to_date:
			raise UserError(_('From date cannot be greater than To date!!'))
		inv_obj = self.env['account.move']
		move_obj = self.env['account.move.line']
		journal_ids = [journal.id for journal in self.journal_id]
		account_ids = [account.id for account in self.account_id]
		if self.od_product == True:
			domain=[('state','in',('open','paid')),('date_invoice','>=',self.from_date),('date_invoice','<=',self.to_date)]
			# inv_datas  = inv_obj.search([('journal_id','in',journal_ids),('state','in',('open','paid')),('date_invoice','>=',self.from_date),('date_invoice','<=',self.to_date)])
			if len(journal_ids) > 0:
				domain.append(('journal_id','in',journal_ids))
			inv_datas  = inv_obj.search(domain)
			inv_datas = sorted(inv_datas, key=lambda k: (datetime.strptime(k.date_invoice, '%Y-%m-%d')))


			return self.generate_excel(inv_datas)
		else:
			if len(account_ids) == 0:
				raise UserError(_('Requires minimum one account!!'))
			domain = [('account_id','in',account_ids),('move_id.state','=','posted'),('date','>=',self.from_date),('date','<=',self.to_date)]
			if len(journal_ids) > 0:
				domain.append(('journal_id','in',journal_ids))
			ir_obj = self.env['ir.model'].search([('model','=','pos.session')])
			if ir_obj:
				move_datas  = move_obj.search(domain)
				move_datas = sorted(move_datas, key=lambda k: (datetime.strptime(k.date, '%Y-%m-%d')))
				return self.generate_excel_pos(move_datas)
			else:
				move_datas  = move_obj.search(domain)
				move_datas = sorted(move_datas, key=lambda k: (datetime.strptime(str(k.date), '%Y-%m-%d')))
				return self.generate_excel_sum(move_datas)

	def generate_excel_pos(self, move_datas):
		company = self.env['res.company'].browse(self.env.user.company_id.id)
		vat_period = datetime.strptime(self.from_date, '%Y-%m-%d').strftime('%d/%m/%Y')+"-"+datetime.strptime(self.to_date,'%Y-%m-%d').strftime('%d/%m/%Y')

		head=defaultdict(list)
		head['add'].append(company.street or "")
		head['add'].append(company.street2 or "")
		head['add'].append(company.state_id.name or "")
		head['add'].append(company.zip or "")
		head['add'].append(company.country_id.name or "")
		head['vat'].append(company.vat or "")
		workbook = xlwt.Workbook(encoding="UTF-8")
		xlwt.add_palette_colour("custom_colour", 0x10)
		workbook.set_colour_RGB(0x10,197,11,11)
		filename='RegisterSummary.xls'
		sheet= workbook.add_sheet('Sales Report',cell_overwrite_ok=True)
		style = xlwt.easyxf('font:name Arial,height 200,colour white;align: horiz center, vert center;pattern: fore_colour custom_colour,pattern solid;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style2 = xlwt.easyxf('font:name Arial,height 200,colour white;align: horiz left, vert center;pattern: fore_colour custom_colour, pattern solid;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style3=xlwt.easyxf('align: horiz center, vert center;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style4=xlwt.easyxf('align: horiz left, vert center;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style5=xlwt.easyxf('align: horiz right, vert center;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style5.num_format_str = "0.00"
		title=['Name of the Company','Name of the Company','Address of the Company','TRN of the Company','VAT Return Period']
		sheet.col(0).width = 256*25
		sheet.col(1).width = 256*63
		sheet.col(2).width = 256*17
		sheet.col(3).width = 256*12
		sheet.col(4).width = 256*12
		sheet.col(5).width = 256*16
		sheet.col(6).width = 256*18
		sheet.col(7).width = 256*18
		sheet.col(8).width = 256*18
		sheet.col(9).width = 256*19
		sheet.col(10).width = 256*14
		sheet.col(11).width = 256*16
		sheet.row(10).height = 256*2
		row = 2
		col =0
		sheet.write(row,col,title[0],style2)
		if company.od_name_arabic==True:
			row=row+1
			sheet.write(row,col,title[1],style2)
		row = row+1
		sheet.write(row,col,title[2],style2)
		row=row+1
		sheet.write(row,col,title[3],style2)
		row=row+1
		sheet.write(row,col,title[4],style2)

		row=2
		col=1
		sheet.write(row,col,company.name,style4)
		if company.od_name_arabic==True:
			row=row+1
			sheet.write(row,col,company.od_name_arabic,style4)
		row=row+1
		sheet.write(row,col,head['add'],style4)
		row=row+1
		sheet.write(row,col,head['vat'],style4)
		row=row+1
		sheet.write(row,col,vat_period,style4)

		header=['Sr.No','Party Name','TRN','Invoice Date','Date of\nSupply','Invoice Number'
				,'Invoice Value\nexcluding VAT AED','Value of VAT\nAED','Total\nincluding VAT AED',
				'VAT Type','Emirates']

		foreign_header=['FYC Code','Value of supply excluding VAT\nin Foreign currency','Value of VAT\nin Foreign currency']# not__included=['Total Value of\nsupplies in AED','Total Value of VAT\nin AED']

		col=0
		row=row+5
		for i in range (0,11):
			sheet.write(row,col,header[i],style)
			col=col+1
		flag=0
		for move_data in move_datas:
			for line in move_data.move_id.invoice_line_ids:
				data = {}
				if move_data.currency_id.name != 'AED' and move_data.currency_id.name != 0:
					flag=1;
		if flag==1:
			sheet.col(12).width = 256*10
			sheet.col(13).width = 256*32
			sheet.col(14).width = 256*21
			for i in range(0,3):
				sheet.write(row,col,foreign_header[i],style)
				col=col+1

		s=0
		inv_total = 0
		tax_total = 0
		grand_total = 0

		for move_data in move_datas:
			data = {}
			temp=0
			d = defaultdict(list)

			pos_amount=0
			data['partner'] = move_data.partner_id.legal_name or move_data.partner_id.name or " "
			data['partner_trn'] = move_data.partner_id.vat or ""
			data['inv_date'] = move_data.date or ''
			data['inv_num']=move_data.move_id.name or " "
			data['tax_amount'] = move_data.debit or -move_data.credit or 0
			data['od_net_amount']=move_data.move_id and move_data.move_id.amount_total or 0
			data['state_id']= move_data.partner_id.state_id.name or ""
			data['fcy']=move_data.currency_id.name or 0
			data['price_subtotal']= move_data.move_id and move_data.move_id.amount_untaxed or 0
			if data['tax_amount']<0:
					data['price_subtotal']=-data['price_subtotal']
			data['inv_date']=datetime.strptime(data['inv_date'], '%Y-%m-%d').strftime('%d/%m/%Y')
			if move_data.move_id:
				for line in move_data.move_id.invoice_line_ids:
					if line.tax_ids:
						for l in line.tax_ids:
							if l.name:
								tax_name=l.name
								d['taxes'].append(tax_name)
					else:
						tax_name=move_data.account_id.name
						d['taxes'].append(tax_name)
			else:
				tax_name=move_data.account_id.name
				d['taxes'].append(tax_name)


			reference=move_data.ref
			pos_obj = False
			if reference:
				pos_obj = self.env['pos.session'].search([('name','=',reference)])
			if pos_obj and len(pos_obj) > 0:
				data['inv_num']=pos_obj.name or move_data.move_id.name or " "
				for pos in pos_obj.statement_ids:
					pos_amount=pos_amount+pos.total_entry_encoding
				data['price_subtotal']=pos_amount or False
				if not d['taxes']:
					if move_data.tax_ids:
						for line in move_data.tax_ids:
							if line.name:
								tax_name=line.name

								d['taxes'].append(tax_name)
					else:
						tax_name=move_data.account_id.name
						d['taxes'].append(tax_name)




			s=s+1
			row=row+1

			col=0
			sheet.write(row,col,s,style3)
			col=col+1
			sheet.write(row,col,data['partner'],style3)
			col=col+1
			sheet.write(row,col,data['partner_trn'],style3)
			col=col+1
			sheet.write(row,col,data['inv_date'],style3)
			col=col+1
			sheet.write(row,col,data['inv_date'],style3)
			col=col+1
			sheet.write(row,col,data['inv_num'],style3)


			if data['fcy'] == 'AED' or data['fcy'] ==0:
				total_amount = data['price_subtotal'] + data['tax_amount']
				col=col+1
				price_subtotal='%.2f'%(data['price_subtotal'])
				tax_amount='%.2f'%(data['tax_amount'])
				total='%.2f'%(total_amount)
				sheet.write(row,col,float(price_subtotal),style5)
				col=col+1
				sheet.write(row,col,float(tax_amount),style5)
				col = col + 1
				sheet.write(row,col,float(total),style5)
				temp=col
				inv_total += data['price_subtotal']
				tax_total += data['tax_amount']
				grand_total += total_amount
			else:
				price = move_data.currency_id.compute(data['price_subtotal'], company.currency_id)
				tax = move_data.currency_id.compute(data['tax_amount'], company.currency_id)
				total_amount = price + tax
				col=col+1
				price_subtotal='%.2f'%(price)
				tax_amount='%.2f'%(tax)
				total='%.2f'%(total_amount)
				sheet.write(row,col,float(price_subtotal),style5)
				col=col+1
				sheet.write(row,col,float(tax_amount),style5)
				temp=col
				inv_total += price
				tax_total += tax
				grand_total +=total_amount
			# col = col + 1
			# sheet.write(row,col,"vbt",style3)
			col=col+1
			if d['taxes'] :
				sheet.write(row,col,d['taxes'][0],style4)
			else:
				sheet.write(row,col," ",style4)

			# sheet.write(row,col,d['taxes'],style4)
			col=col+1
			sheet.write(row,col,data['state_id'],style3)
			col=col+1
			# sheet.write(row,col,"",style3)
			if data['fcy'] != 'AED' and data['fcy'] != 0:
				total_amount = data['price_subtotal'] + data['tax_amount']
				col=col+1
				sheet.write(row,col,data['fcy'],style3)
				col=col+1
				price_subtotal='%.2f'%(data['price_subtotal'])
				tax_amount='%.2f'%(data['tax_amount'])
				total='%.2f'%(total_amount)
				sheet.write(row,col,float(price_subtotal),style5)
				col=col+1
				sheet.write(row,col,float(tax_amount),style5)
				col = col + 1
				sheet.write(row,col,float(total),style5)


		row = row + 1
		col = 0
		sheet.write(row,col+5,'TOTAL ',style3)
		inv_total='%.2f'%(inv_total)
		tax_total='%.2f'%(tax_total)
		grand_total='%.2f'%(grand_total)
		sheet.write(row,col+6,float(inv_total),style5)
		sheet.write(row,col+7,float(tax_total),style5)
		sheet.write(row,col+8,float(grand_total),style5)

		fp = StringIO()
		workbook.save(fp)
		excel_file = base64.encodestring(fp.getvalue())
		self.excel_file = excel_file
		self.file_name =filename
		fp.close()
		return {
			  'view_type': 'form',
			  "view_mode": 'form',
			  'res_model': 'orchid.tax.register',
			  'res_id': self.id,
			  'type': 'ir.actions.act_window',
			  'target': 'new'
			  }

	def generate_excel(self, inv_datas):
		company = self.env['res.company'].browse(self.env.user.company_id.id)
		vat_period = datetime.strptime(self.from_date, '%Y-%m-%d').strftime('%d/%m/%Y')+"-"+datetime.strptime(self.to_date,'%Y-%m-%d').strftime('%d/%m/%Y')
		head=defaultdict(list)
		head['add'].append(company.street or "")
		head['add'].append(company.street2 or "")
		head['add'].append(company.state_id.name or "")
		head['add'].append(company.zip or "")
		head['add'].append(company.country_id.name or "")
		head['vat'].append(company.vat or "")

		workbook= xlwt.Workbook(encoding="UTF-8")
		xlwt.add_palette_colour("custom_colour", 0x10)
		workbook.set_colour_RGB(0x10,197,11,11)
		filename='Register.xls'
		sheet= workbook.add_sheet('Sales Report',cell_overwrite_ok=True)
		style = xlwt.easyxf('font:name Arial,height 200,colour white;align: horiz center, vert center;pattern: fore_colour custom_colour,pattern solid;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style2 = xlwt.easyxf('font:name Arial,height 200,colour white;align: horiz left, vert center;pattern: fore_colour custom_colour, pattern solid;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style3=xlwt.easyxf('align: horiz center, vert center;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style4=xlwt.easyxf('align: horiz left, vert center;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style5=xlwt.easyxf('align: horiz right, vert center;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style5.num_format_str = "0.00"
		title=['Name of the Company','Name of the Company','Address of the Company','TRN of the Company','VAT Return Period']
		sheet.col(0).width = 256*25
		sheet.col(1).width = 256*63
		sheet.col(2).width = 256*17
		sheet.col(3).width = 256*12
		sheet.col(4).width = 256*12
		sheet.col(5).width = 256*16
		sheet.col(6).width = 256*60
		sheet.col(7).width = 256*18
		sheet.col(8).width = 256*18
		sheet.col(9).width = 256*19
		sheet.col(10).width = 256*14
		sheet.col(11).width = 256*16
		sheet.row(10).height = 256*2
		row = 2
		col =0
		sheet.write(row,col,title[0],style2)
		if company.od_name_arabic==True:
			row=row+1
			sheet.write(row,col,title[1],style2)
		row=row+1
		sheet.write(row,col,title[2],style2)

		row=row+1
		sheet.write(row,col,title[3],style2)
		row=row+1
		sheet.write(row,col,title[4],style2)

		row=2
		col=1
		sheet.write(row,col,company.name,style4)
		if company.od_name_arabic==True:
			row=row+1
			sheet.write(row,col,company.od_name_arabic,style4)

		row=row+1
		sheet.write(row,col,head['add'],style4)
		row=row+1
		sheet.write(row,col,head['vat'],style4)
		row=row+1
		sheet.write(row,col,vat_period,style4)

		header=['Sr.No','Party Name','TRN','Invoice Date','Date of\nSupply','Invoice Number'
				,'Product\ndescription','Invoice Value\nexcluding VAT AED','Value of VAT\nAED',
				'Total\nincluding VAT AED','VAT Type','Emirates']

		foreign_header=['FYC Code','Value of supply excluding VAT\nin Foreign currency','Value of VAT\nin Foreign currency']# not__included=['Total Value of\nsupplies in AED','Total Value of VAT\nin AED']

		col=0
		row=row+5
		for i in range (0,12):
			sheet.write(row,col,header[i],style)
			col=col+1
		flag=0
		for invoice_data in inv_datas:
			for line in invoice_data.invoice_line_ids:
				data = {}
				if invoice_data.currency_id.name != 'AED':
					flag=1;
		if flag==1:
			sheet.col(12).width = 256*10
			sheet.col(13).width = 256*32
			sheet.col(14).width = 256*21
			for i in range(0,3):
				sheet.write(row,col,foreign_header[i],style)
				col=col+1

		s=0
		for invoice_data in inv_datas:
			for line in invoice_data.invoice_line_ids:
				data = {}
				temp=0
				d = defaultdict(list)
				data['partner'] = invoice_data.partner_id.legal_name or invoice_data.partner_id.name or " "
				data['partner_trn'] = invoice_data.partner_id.vat or ""
				data['inv_date'] = invoice_data.date_invoice or ''
				data['inv_num'] = invoice_data.number or ''
				data['pro_desc'] = line.name or ""
				data['price_subtotal']=line.price_subtotal or 0
				data['tax_amount'] = line.od_tax_amount or 0
				data['od_net_amount']=line.od_net_amount or 0
				data['state_id']= line.partner_id.state_id.name or ""
				data['fcy']=invoice_data.currency_id.name
				if data['tax_amount']<0:
					data['price_subtotal']=-data['price_subtotal']
				data['inv_date']=datetime.strptime(data['inv_date'], '%Y-%m-%d').strftime('%d/%m/%Y')

				s=s+1
				row=row+1

				col=0
				sheet.write(row,col,s,style3)
				col=col+1
				sheet.write(row,col,data['partner'],style3)
				col=col+1
				sheet.write(row,col,data['partner_trn'],style3)
				col=col+1
				sheet.write(row,col,data['inv_date'],style3)
				col=col+1
				sheet.write(row,col,data['inv_date'],style3)
				col=col+1
				sheet.write(row,col,data['inv_num'],style3)
				col=col+1
				sheet.write(row,col,data['pro_desc'],style4)

				if data['fcy'] == 'AED':
					total_amount = float(data['price_subtotal'] or 0) + float(data['tax_amount'] or 0)
					col=col+1
					price_subtotal='%.2f'%(data['price_subtotal'] or 0)
					tax_amount='%.2f'%(data['tax_amount'] or 0)
					total='%.2f'%(total_amount)
					sheet.write(row,col,float(price_subtotal),style5)
					col=col+1
					sheet.write(row,col,float(tax_amount),style5)
					col = col + 1
					sheet.write(row,col,float(total),style5)
					temp=col
				else:
					price = invoice_data.currency_id.compute(line.price_subtotal, company.currency_id)
					tax=invoice_data.currency_id.compute(line.od_tax_amount, company.currency_id)
					total_amount = price + tax
					col=col+1
					price_subtotal='%.2f'%(price)
					tax_amount='%.2f'%(tax)
					total='%.2f'%(total_amount)
					sheet.write(row,col,float(price_subtotal),style5)
					col=col+1
					sheet.write(row,col,float(tax_amount),style5)
					col = col + 1
					sheet.write(row,col,float(total),style5)
					temp=col

				for l in line.tax_ids:
					tax_name=l.name or move_data.account_id
					d['taxes'].append(tax_name)

				col=col+1
				sheet.write(row,col,d['taxes'],style4)
				col=col+1
				sheet.write(row,col,data['state_id'],style3)
				# col=col+1
				# sheet.write(row,col,"",style3)
				if data['fcy'] != 'AED':
					total_amount = data['price_subtotal'] + data['tax_amount']
					col=col+1
					sheet.write(row,col,data['fcy'],style3)
					col=col+1
					price_subtotal='%.2f'%(data['price_subtotal'])
					tax_amount='%.2f'%(data['tax_amount'])
					total='%.2f'%(total_amount)
					sheet.write(row,col,float(price_subtotal),style5)
					col=col+1
					sheet.write(row,col,float(tax_amount),style5)
					col = col + 1
					sheet.write(row,col,float(total),style5)


		fp = StringIO()
		workbook.save(fp)
		excel_file = base64.encodestring(fp.getvalue())
		self.excel_file = excel_file
		self.file_name =filename
		fp.close()
		return {
			  'view_type': 'form',
			  "view_mode": 'form',
			  'res_model': 'orchid.tax.register',
			  'res_id': self.id,
			  'type': 'ir.actions.act_window',
			  'target': 'new'
			  }

	def generate_excel_sum(self, move_datas):
		company = self.env['res.company'].browse(self.env.user.company_id.id)
		vat_period = datetime.strptime(str(self.from_date), '%Y-%m-%d').strftime('%d/%m/%Y')+"-"+datetime.strptime(str(self.to_date),'%Y-%m-%d').strftime('%d/%m/%Y')

		head=defaultdict(list)
		head['add'].append(company.street or "")
		head['add'].append(company.street2 or "")
		head['add'].append(company.state_id.name or "")
		head['add'].append(company.zip or "")
		head['add'].append(company.country_id.name or "")
		head['vat'].append(company.vat or "")

		workbook = xlwt.Workbook(encoding="UTF-8")
		xlwt.add_palette_colour("custom_colour", 0x10)
		workbook.set_colour_RGB(0x10,197,11,11)
		filename='RegisterSummary.xls'
		sheet= workbook.add_sheet('Sales Report',cell_overwrite_ok=True)
		style = xlwt.easyxf('font:name Arial,height 200,colour white;align: horiz center, vert center;pattern: fore_colour custom_colour,pattern solid;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style2 = xlwt.easyxf('font:name Arial,height 200,colour white;align: horiz left, vert center;pattern: fore_colour custom_colour, pattern solid;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style3=xlwt.easyxf('align: horiz center, vert center;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style4=xlwt.easyxf('align: horiz left, vert center;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style5=xlwt.easyxf('align: horiz right, vert center;borders:top_color black, bottom_color black, right_color black, left_color black, top thin,right thin,bottom thin,left thin;')
		style5.num_format_str = "0.00"

		title=['Name of the Company','Name of the Company','Address of the Company','TRN of the Company','VAT Return Period']
		sheet.col(0).width = 256*25
		sheet.col(1).width = 256*63
		sheet.col(2).width = 256*17
		sheet.col(3).width = 256*32
		sheet.col(4).width = 256*12
		sheet.col(5).width = 256*12
		sheet.col(6).width = 256*60
		sheet.col(7).width = 256*19
		sheet.col(8).width = 256*18
		sheet.col(9).width = 256*18
		sheet.col(10).width = 256*19
		sheet.col(11).width = 256*14
		sheet.col(12).width = 256*16
		sheet.row(10).height = 256*2
		row = 2
		col =0
		sheet.write(row,col,title[0],style2)
		if company.od_name_arabic==True:
			row=row+1
			sheet.write(row,col,title[1],style2)
		row = row+1
		sheet.write(row,col,title[2],style2)
		row=row+1
		sheet.write(row,col,title[3],style2)
		row=row+1
		sheet.write(row,col,title[4],style2)



		row=2
		col=1
		sheet.write(row,col,company.name,style4)
		if company.od_name_arabic==True:
			row=row+1
			sheet.write(row,col,company.od_name_arabic,style4)
		row=row+1
		sheet.write(row,col,head['add'],style4)
		row=row+1
		sheet.write(row,col,head['vat'],style4)
		row=row+1
		sheet.write(row,col,vat_period,style4)

		header=['Sr.No','Party Name','TRN','Sales Person','Invoice Date','Date of\nSupply','Invoice Number',
				'Reference','Invoice Value\nexcluding VAT AED','Value of VAT\nAED',
				'Total\nincluding VAT AED','VAT Type','Emirates']

		foreign_header=['FYC Code','Value of supply excluding VAT\nin Foreign currency','Value of VAT\nin Foreign currency']# not__included=['Total Value of\nsupplies in AED','Total Value of VAT\nin AED']

		col=0
		row=row+5
		for i in range (0,13):
			sheet.write(row,col,header[i],style)
			col=col+1
		flag=0
		for move_data in move_datas:
			for line in move_data.move_id.invoice_line_ids:
				data = {}
				if move_data.currency_id.name != 'AED' and move_data.currency_id.name != 0:
					flag=1;
		if flag==1:
			sheet.col(13).width = 256*10
			sheet.col(14).width = 256*32
			sheet.col(15).width = 256*21
			for i in range(0,3):
				sheet.write(row,col,foreign_header[i],style)
				col=col+1

		s=0
		inv_total = 0
		tax_total = 0
		grand_total = 0

		for move_data in move_datas:
			data = {}
			temp = 0
			d = defaultdict(list)

			data['partner'] = move_data.partner_id.name or " "
			data['partner_trn'] = move_data.partner_id.vat or ""
			data['sales_person'] = move_data.move_id and move_data.move_id.user_id and move_data.move_id.user_id.name or ""
			data['inv_date'] = move_data.date or ''
			data['inv_num'] = move_data.move_id.name or ''
			data['inv_ref'] = move_data.name or ''

			# Fixed for Odoo 18
			if move_data.move_id and move_data.move_id.is_invoice():
				data['price_subtotal'] = move_data.move_id.amount_untaxed or 0
			else:
				data['price_subtotal'] = abs(move_data.balance) or 0

			data['tax_amount'] = move_data.debit or -move_data.credit or 0
			data['od_net_amount'] = move_data.move_id and move_data.move_id.amount_total or 0
			data['state_id'] = move_data.partner_id.state_id.name or " "
			data['fcy'] = move_data.currency_id.name or 0
			if data['tax_amount']<0:
				data['price_subtotal']=-data['price_subtotal']
			data['inv_date']=datetime.strptime(str(data['inv_date']), '%Y-%m-%d').strftime('%d/%m/%Y')
			s=s+1
			row=row+1

			col=0
			sheet.write(row,col,s,style3)
			col=col+1
			sheet.write(row,col,data['partner'],style3)
			col=col+1
			sheet.write(row,col,data['partner_trn'],style3)
			col=col+1
			sheet.write(row,col,data['sales_person'],style3)
			col=col+1
			sheet.write(row,col,data['inv_date'],style3)
			col=col+1
			sheet.write(row,col,data['inv_date'],style3)
			col=col+1
			sheet.write(row,col,data['inv_num'],style3)
			col=col+1
			sheet.write(row,col,data['inv_ref'],style3)

			if data['fcy'] == 'AED' or data['fcy'] ==0:
				total_amount = float(data['price_subtotal'] or 0) + float(data['tax_amount'] or 0)
				col=col+1
				price_subtotal='%.2f'%(data['price_subtotal'])
				tax_amount='%.2f'%(data['tax_amount'])
				total='%.2f'%(total_amount)
				sheet.write(row,col,float(price_subtotal),style5)
				col=col+1
				sheet.write(row,col,float(tax_amount),style5)
				col = col + 1
				sheet.write(row,col,float(total),style5)
				temp=col
				inv_total += data['price_subtotal']
				tax_total += data['tax_amount']
				grand_total += total_amount
			else:
				price = move_data.currency_id.compute(data['price_subtotal'], company.currency_id)
				tax = move_data.currency_id.compute(data['tax_amount'], company.currency_id)
				total_amount = price + tax
				col=col+1

				price_subtotal='%.2f'%(price)
				tax_amount='%.2f'%(tax)
				total='%.2f'%(total_amount)
				sheet.write(row,col,float(price_subtotal),style5)
				col=col+1
				sheet.write(row,col,float(tax_amount),style5)
				col = col + 1
				sheet.write(row,col,float(total),style5)
				temp=col
				inv_total += price
				tax_total += tax
				grand_total += total_amount

			for line in move_data.move_id.invoice_line_ids:
				for l in line.tax_ids:
					tax_name=l.name or move_data.account_id
					d['taxes'].append(l.name or ' ')
			col=col+1
			if d['taxes'] :
				sheet.write(row,col,d['taxes'][0],style4)
			else:
				sheet.write(row,col," ",style4)
			col=col+1
			sheet.write(row,col,data['state_id'],style3)
			# col=col+1
			# sheet.write(row,col,"",style3)
			if data['fcy'] != 'AED' and data['fcy'] != 0:
				total_amount = data['price_subtotal'] + data['tax_amount']
				col=col+1
				sheet.write(row,col,data['fcy'],style3)
				col=col+1
				price_subtotal='%.2f'%(data['price_subtotal'])
				tax_amount='%.2f'%(data['tax_amount'])
				total='%.2f'%(total_amount)
				sheet.write(row,col,float(price_subtotal),style5)
				col=col+1
				sheet.write(row,col,float(tax_amount),style5)
				col = col + 1
				sheet.write(row,col,float(total),style5)

		row = row + 1
		col = 0
		sheet.write(row,col+7,'TOTAL ',style3)
		inv_total='%.2f'%(inv_total)
		tax_total='%.2f'%(tax_total)
		grand_total='%.2f'%(grand_total)
		sheet.write(row,col+8,float(inv_total),style5)
		sheet.write(row,col+9,float(tax_total),style5)
		sheet.write(row,col+10,float(grand_total),style5)

		fp = BytesIO()
		workbook.save(fp)
		excel_file =  base64.b64encode(fp.getvalue())
		self.excel_file = excel_file
		self.file_name =filename
		fp.close()
		return {
			  'view_type': 'form',
			  "view_mode": 'form',
			  'res_model': 'orchid.tax.register',
			  'res_id': self.id,
			  'type': 'ir.actions.act_window',
			  'target': 'new'
			  }





